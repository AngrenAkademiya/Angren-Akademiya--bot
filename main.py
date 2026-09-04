import os
import logging
from datetime import datetime
import asyncio
import json
import random
from aiogram import Bot, Dispatcher, types, F
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext
from aiogram.utils.keyboard import ReplyKeyboardBuilder, InlineKeyboardBuilder
from aiohttp import web, ClientSession
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.page import PageMargins
from aiogram.types import FSInputFile
import gspread
from google.oauth2.service_account import Credentials
from PIL import Image, ImageDraw, ImageFont
import qrcode
logging.basicConfig(level=logging.INFO)

API_TOKEN = os.getenv("BOT_TOKEN")
if not API_TOKEN:
    raise ValueError("Xatolik: BOT_TOKEN topilmadi! Render sozlamalarini tekshiring.")

bot = Bot(token=API_TOKEN)
dp = Dispatcher(storage=MemoryStorage())

EXCEL_FILE = "students.xlsx"

CHANNEL_USERNAME = "@AngrenAkademiya"

async def check_subscription(user_id: int) -> bool:
    try:
        member = await bot.get_chat_member(chat_id=CHANNEL_USERNAME, user_id=user_id)
        return member.status not in ("left", "kicked")
    except Exception:
        logging.exception("Kanalga a'zolikni tekshirishda xato:")
        return False

def get_subscribe_keyboard():
    kb = InlineKeyboardBuilder()
    kb.button(text="📍 Kanalga o'tish", url="https://t.me/AngrenAkademiya")
    kb.button(text="✅ A'zo bo'ldim, tekshirish", callback_data="check_sub")
    kb.adjust(1)
    return kb.as_markup()

def generate_certificate(full_name: str, user_id: int) -> str:
    template = Image.open("certificate_template.png").convert("RGB")
    draw = ImageDraw.Draw(template)
    font = ImageFont.truetype(
        "/usr/share/fonts/truetype/liberation/LiberationSerif-BoldItalic.ttf", 42
    )
    text = full_name.upper()
    bbox = draw.textbbox((0, 0), text, font=font)
    w = bbox[2] - bbox[0]
    x = (template.width - w) / 2
    y = 340  # 470 dan 520 ga tushirildi — chiziq ustiga to'g'ri kelishi uchun
    draw.text((x, y), text, font=font, fill=(255, 255, 255))

    date_text = datetime.now().strftime("%d.%m.%Y")
    draw.text(
        (100, template.height - 60), f"Sana: {date_text}",
        font=ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 24),
        fill=(180, 180, 180)
    )

    qr_link = f"https://t.me/AngrenAkademiya_bot?start=cabinet_{user_id}"
    qr_img = qrcode.make(qr_link).resize((110, 110))
    template.paste(qr_img, (template.width - 200, 330))

    out_path = f"cert_{user_id}.png"
    template.save(out_path)
    return out_path

DIPLOMA_TIERS = [
    (100, "mutlaq_g'olib.png", "MUTLAQ G'OLIB"),
    (95, "a'lo_darajali.png", "A'LO DARAJALI O'QUVCHI"),
    (90, "faol_va_iqtidorli.png", "FAOL VA IQTIDORLI O'QUVCHI"),
    (85, "iqtidorli.png", "IQTIDORLI O'QUVCHI"),
    (80, "bilimdon.png",  "BILIMDON O'QUVCHI"),
    (70, "bilimga_intiluvchi.png", "BILIMGA INTILUVCHI O'QUVCHI"),
]


def get_diploma_tier(percent: int):
    for min_percent, filename, title in DIPLOMA_TIERS:
        if percent >= min_percent:
            return filename, title
    return None, None


def get_student_level(percent: float) -> str:
    if percent == 100:
        return "🏆 Mutlaq a'lochi"
    elif 86 <= percent <= 99:
        return "🥇 A'lochi o'quvchi"
    elif 70 <= percent <= 85:
        return "🥈 Yaxshi bilimli"
    elif 56 <= percent <= 69:
        return "🥉 O'rtacha bilimli"
    else:
        return "⚠️ Past o'zlashtiruvchi"

def find_full_name_by_id(user_id: int) -> str:
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[10] == user_id:
                return row[2]
    except Exception:
        logging.exception("Ism qidirishda xato:")
    return None


def find_student_by_id(user_id: int):
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[10] == user_id:
                voucher = row[11] if len(row) > 11 and row[11] is not None else 30000
                referrals = row[12] if len(row) > 12 and row[12] is not None else 0
                return {
                    "name": row[2],
                    "parent_phone": row[4],
                    "school": row[5],
                    "grade": row[6],
                    "voucher": voucher,
                    "referrals": referrals,
                }
    except Exception:
        logging.exception("O'quvchi ma'lumotini qidirishda xato:")
    return None


def generate_diploma(full_name: str, subject: str, percent: int, user_id: int) -> str:
    filename, title = get_diploma_tier(percent)
    if not filename:
        return None

    template = Image.open(filename).convert("RGB")
    draw = ImageDraw.Draw(template)

    name_font = ImageFont.truetype(
        "/usr/share/fonts/truetype/liberation/LiberationSerif-BoldItalic.ttf", 36
    )
    subject_font = ImageFont.truetype(
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 26
    )

    name_text = full_name.upper()
    bbox = draw.textbbox((0, 0), name_text, font=name_font)
    w = bbox[2] - bbox[0]
    x = (template.width - w) / 2
    y = int(template.height * 0.44)
    draw.text((x, y), name_text, font=name_font, fill=(20, 110, 60))

    bbox2 = draw.textbbox((0, 0), subject, font=subject_font)
    w2 = bbox2[2] - bbox2[0]
    x2 = (template.width - w2) / 2
    y2 = int(template.height * 0.575)
    draw.text((x2, y2), subject, font=subject_font, fill=(184, 134, 11))

    out_path = f"diploma_{user_id}.png"
    template.save(out_path)
    return out_path
def save_to_excel(data, user_id):
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "O'quvchilar"
        ws.append(["№", "Sana", "Ism Familiya", "Tel Raqam", "Ota-ona Tel", "Maktab", "Sinf", "Filial", "Smena", "Kurslar", "ID", "Voucher", "Taklif qilganlar"])
        wb.save(EXCEL_FILE)

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    sana = datetime.now().strftime("%d.%m.%Y %H:%M")

    courses_list = data.get("selected_courses", [])
    courses_string = "\n".join(f"• {c.replace(chr(10), ' ')}" for c in courses_list)

    tartib_raqam = ws.max_row

    ws.append([
        tartib_raqam,
        sana,
        data.get("name"),
        data.get("phone"),
        data.get("parent_phone"),
        data.get("school"),
        data.get("grade"),
        data.get("filial"),
        data.get("time_pref"),
        courses_string,
        user_id,
        30000,
        0
    ])

    try:
        col_widths = {1: 5, 2: 16, 3: 22, 4: 14, 5: 14, 6: 10, 7: 6, 8: 12, 9: 12, 10: 15, 11: 15, 12: 12, 13: 16}
        for col_num, width in col_widths.items():
            col_letter = openpyxl.utils.get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = width

        for row in ws.iter_rows(min_row=1):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")

        for row in ws.iter_rows(min_row=2):
            max_lines = 1
            for cell in row:
                if cell.value:
                    lines_count = str(cell.value).count("\n") + 1
                    if lines_count > max_lines:
                        max_lines = lines_count
            ws.row_dimensions[row[0].row].height = max(max_lines * 15, 20)

        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)

    except Exception as e:
        logging.warning(f"Excel formatlashda xato (malumot yozildi): {e}")

    wb.save(EXCEL_FILE)


def _write_to_google_sheets_sync(data, user_id):
    now = datetime.now()
    sana = now.strftime("%d.%m.%Y %H:%M")
    bugun = now.strftime("%d.%m.%Y")

    courses_list = data.get("selected_courses", [])
    courses_string = ", ".join([c.replace('\n', ' ') for c in courses_list])

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]

    creds_json = os.getenv("GOOGLE_CREDS")

    if creds_json:
        creds_data = json.loads(creds_json)
        creds = Credentials.from_service_account_info(creds_data, scopes=scopes)
    else:
        creds = Credentials.from_service_account_file("credentials.json", scopes=scopes)

    client = gspread.authorize(creds)

    sheet_url = os.getenv("GOOGLE_SHEET_URL")
    spreadsheet_id = os.getenv("SPREADSHEET_ID")

    if sheet_url:
        spreadsheet = client.open_by_url(sheet_url)
    elif spreadsheet_id:
        spreadsheet = client.open_by_key(spreadsheet_id)
    else:
        spreadsheet = client.open_by_url(
            "https://docs.google.com/spreadsheets/d/1aXoL-TeP0Oh62u1kfgPyzyRsNjOdqGkovJmFutYlUn0/edit"
        )

    try:
        sheet = spreadsheet.worksheet(bugun)
    except Exception:
        sheet = spreadsheet.add_worksheet(title=bugun, rows=1000, cols=13)
        sheet.append_row([
            "№", "Sana", "Ism Familiya", "Tel Raqam",
            "Ota-ona Tel", "Maktab", "Sinf", "Filial", "Smena", "Kurslar", "ID", "Voucher", "Taklif qilganlar"
        ])

    all_rows = sheet.get_all_values()
    tartib_raqam = len(all_rows)

    sheet.append_row([
        tartib_raqam,
        sana,
        data.get("name"),
        data.get("phone"),
        data.get("parent_phone"),
        data.get("school"),
        data.get("grade"),
        data.get("filial"),
        data.get("time_pref"),
        courses_string,
        user_id,
        30000,
        0
    ])


async def save_to_google_sheets(data, user_id):
    try:
        await asyncio.to_thread(_write_to_google_sheets_sync, data, user_id)
        logging.info("Ma'lumotlar Google Sheets'ga muvaffaqiyatli yozildi!")
    except Exception as e:
        logging.exception("Google Sheets yozishda xato:")
        admin_id = os.getenv("ADMIN_ID")
        if admin_id:
            try:
                await bot.send_message(
                    int(admin_id),
                    f"Google Sheets xato: {type(e).__name__}: {e}"
                )
            except Exception:
                pass


def _write_test_result_sync(user_id, subject, grade, percent):
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]
    creds_json = os.getenv("GOOGLE_CREDS")
    if creds_json:
        creds_data = json.loads(creds_json)
        creds = Credentials.from_service_account_info(creds_data, scopes=scopes)
    else:
        creds = Credentials.from_service_account_file("credentials.json", scopes=scopes)
    client = gspread.authorize(creds)

    sheet_url = os.getenv("GOOGLE_SHEET_URL")
    spreadsheet_id = os.getenv("SPREADSHEET_ID")
    if sheet_url:
        spreadsheet = client.open_by_url(sheet_url)
    elif spreadsheet_id:
        spreadsheet = client.open_by_key(spreadsheet_id)
    else:
        spreadsheet = client.open_by_url(
            "https://docs.google.com/spreadsheets/d/1aXoL-TeP0Oh62u1kfgPyzyRsNjOdqGkovJmFutYlUn0/edit"
        )

    try:
        sheet = spreadsheet.worksheet("Test Natijalari")
    except Exception:
        sheet = spreadsheet.add_worksheet(title="Test Natijalari", rows=2000, cols=8)
        sheet.append_row(["Sana", "O'quvchi", "Maktab", "Sinf", "Fan", "Natija (%)", "Daraja", "ID"])

    student = find_student_by_id(user_id) or {}
    name = student.get("name") or "Noma'lum"
    school = student.get("school") or "—"
    sana = datetime.now().strftime("%d.%m.%Y %H:%M")
    daraja = get_student_level(percent)

    sheet.append_row([sana, name, school, grade, subject, percent, daraja, user_id])


async def save_test_result(user_id, subject, grade, percent):
    try:
        await asyncio.to_thread(_write_test_result_sync, user_id, subject, grade, percent)
        logging.info("Test natijasi 'Test Natijalari' jadvaliga yozildi!")
    except Exception:
        logging.exception("Test natijasini yozishda xato:")


async def notify_admin_test_result(user_id, subject, grade, score, total, percent):
    admin_id = os.getenv("ADMIN_ID")
    if not admin_id:
        return
    try:
        student = await asyncio.to_thread(find_student_by_id, user_id) or {}
        name = student.get("name") or "Noma'lum"
        school = student.get("school") or "—"
        parent_phone = student.get("parent_phone") or "—"
        daraja = get_student_level(percent)
        sana = datetime.now().strftime("%Y-%m-%d %H:%M")

        text = (
            f"📊 TEST NATIJASI HISOBOTI\n\n"
            f"👤 O'quvchi: {name}\n"
            f"🏫 Maktab/Sinf: {school} | {grade}-sinf\n"
            f"📚 Fan: {subject}\n"
            f"🎯 Natija: {score}/{total} ball ({percent}%)\n"
            f"🎖 Daraja: {daraja}\n"
            f"📞 Ota-ona tel: {parent_phone}\n"
            f"⏰ Vaqt: {sana}"
        )
        await bot.send_message(int(admin_id), text)
    except Exception:
        logging.exception("Admin xabarnomasini yuborishda xato:")


class RoleRequest(StatesGroup):
    maktab = State()
    role = State()
    sinf = State()


class Registration(StatesGroup):
    name = State()
    phone = State()
    parent_phone = State()
    school = State()
    grade = State()
    grade_letter = State()
    filial = State()
    subjects = State()
    time_pref = State()


AVAILABLE_FILIALS = ["Angren", "Ohangaron"]

ANGREN_SCHOOLS = [f"{i}-maktab" for i in range(1, 29)]
AVAILABLE_TIMES = ["Ertalabki", "Kunduzgi", "Kechki"]
GRADE_LETTERS = ["A", "B", "V", "G", "D", "E", "J", "Z", "I", "K", "L"]

COURSE_CATEGORIES = {
    "🩺 Shifokorlik yo'nalishi (5 ta fan bir joyda)": [
        "1️⃣ Kimyo - Milliy va xalqaro sertifikat",
        "2️⃣ Biologiya - Milliy va xalqaro sertifikat",
        "3️⃣ Majburiy fanlar (Matematika, Tarix, Ona tili) - sertifikatlar"
    ],
    "📚 Majburiy blok (Hamma uchun)": [
        "📌 Majburiy fanlar (Matematika, Tarix, Ona tili) - sertifikatlar"
    ],
    "⚡ Qisqa muddatli tayyorlov kurslari": [
        "🚀 Tezlashtirilgan (Intensiv) tayyorlov",
        "📝 Imtihon oldi takrorlash kursi"
    ],
    "🌱 Past o'zlashtiruvchilar bilan ishlash": [
        "🎯 Maxsus yakka tartibdagi (individual) dastur"
    ],
    "📐 Aniq fanlar & IT": [
        "🧮 Matematika - Milliy va xalqaro sertifikat",
        "📐 Fizika - Milliy va xalqaro sertifikat",
        "💻 IT - Milliy va xalqaro sertifikat"
    ],
    "🌍 Tillar va Gumanitar": [
        "🇬🇧 Ingliz tili - IELTS / CEFR",
        "🇷🇺 Rus tili - Milliy va xalqaro sertifikat",
        "📖 Ona tili va adabiyot - Milliy sertifikat",
        "📜 Tarix - Milliy sertifikat",
        "⚖️ Huquq - Milliy sertifikat"
    ],
    "🏫 Prezident maktabi va maxsus maktablarga tayyorlov": [
        "⭐ Prezident maktablariga tayyorlov",
        "📐 Al-Xorazmiy maktabiga tayyorlov",
        "🧸 Maktabga tayyorlov (Pochemuchka)"
    ]
}
COURSE_CATEGORY_LIST = list(COURSE_CATEGORIES.items())

class TestQuiz(StatesGroup):
    subject = State()
    grade = State()
    question = State()

MATH_QUESTIONS = {
    1: [
        {"q": "2+3=?", "options": ["4","5","6","7"], "correct": 1},
        {"q": "7-2=?", "options": ["4","5","6","3"], "correct": 1},
        {"q": "1+1=?", "options": ["1","2","3","4"], "correct": 1},
        {"q": "9-4=?", "options": ["3","4","5","6"], "correct": 2},
        {"q": "5+4=?", "options": ["7","8","9","10"], "correct": 2},
        {"q": "10-6=?", "options": ["3","4","5","6"], "correct": 1},
        {"q": "6+3=?", "options": ["8","9","10","7"], "correct": 1},
        {"q": "8-5=?", "options": ["2","3","4","5"], "correct": 1},
        {"q": "4+4+1=?", "options": ["8","9","10","7"], "correct": 1},
        {"q": "10-3-2=?", "options": ["4","5","6","3"], "correct": 1},
    ],
    2: [
        {"q": "12+7=?", "options": ["18","19","20","17"], "correct": 1},
        {"q": "15-8=?", "options": ["6","7","8","9"], "correct": 1},
        {"q": "6×2=?", "options": ["10","12","14","8"], "correct": 1},
        {"q": "20-9=?", "options": ["10","11","12","9"], "correct": 1},
        {"q": "9+9=?", "options": ["16","17","18","19"], "correct": 2},
        {"q": "4×5=?", "options": ["18","20","22","16"], "correct": 1},
        {"q": "36-18=?", "options": ["16","17","18","19"], "correct": 2},
        {"q": "7×3=?", "options": ["19","20","21","22"], "correct": 2},
        {"q": "45+27=?", "options": ["70","71","72","73"], "correct": 2},
        {"q": "60-24=?", "options": ["34","35","36","37"], "correct": 2},
    ],
    3: [
        {"q": "8×7=?", "options": ["54","55","56","57"], "correct": 2},
        {"q": "63÷9=?", "options": ["6","7","8","9"], "correct": 1},
        {"q": "100-45=?", "options": ["54","55","56","57"], "correct": 1},
        {"q": "9×6=?", "options": ["52","53","54","55"], "correct": 2},
        {"q": "72÷8=?", "options": ["7","8","9","10"], "correct": 2},
        {"q": "125+237=?", "options": ["360","361","362","363"], "correct": 2},
        {"q": "84÷7=?", "options": ["11","12","13","14"], "correct": 1},
        {"q": "15×3=?", "options": ["43","44","45","46"], "correct": 2},
        {"q": "500-268=?", "options": ["230","231","232","233"], "correct": 2},
        {"q": "96÷8÷2=?", "options": ["5","6","7","8"], "correct": 1},
    ],
    4: [
        {"q": "234+567=?", "options": ["799","800","801","802"], "correct": 2},
        {"q": "12×12=?", "options": ["142","143","144","145"], "correct": 2},
        {"q": "900÷30=?", "options": ["28","29","30","31"], "correct": 2},
        {"q": "1000-456=?", "options": ["543","544","545","546"], "correct": 1},
        {"q": "15×20=?", "options": ["280","290","300","310"], "correct": 2},
        {"q": "3/4 + 1/4 = ?", "options": ["1/2","3/4","1","5/4"], "correct": 2},
        {"q": "6.5+3.2=?", "options": ["9.5","9.6","9.7","9.8"], "correct": 2},
        {"q": "144÷12=?", "options": ["11","12","13","14"], "correct": 1},
        {"q": "Tomoni 6 sm kvadrat perimetri?", "options": ["18","20","22","24"], "correct": 3},
        {"q": "2500÷25=?", "options": ["90","95","100","105"], "correct": 2},
    ],
    5: [
        {"q": "3/5 + 1/5 = ?", "options": ["2/5","4/5","3/10","1"], "correct": 1},
        {"q": "25% dan 200 necha?", "options": ["40","45","50","55"], "correct": 2},
        {"q": "7² = ?", "options": ["14","42","49","56"], "correct": 2},
        {"q": "0.75 ni foizga aylantiring", "options": ["65%","70%","75%","80%"], "correct": 2},
        {"q": "18×15=?", "options": ["260","270","280","290"], "correct": 1},
        {"q": "To'g'ri to'rtburchak yuzi: 8×5=?", "options": ["35","38","40","42"], "correct": 2},
        {"q": "-5+8=?", "options": ["2","3","4","13"], "correct": 1},
        {"q": "2/3 ning 1/2 qismi?", "options": ["1/6","1/3","2/6","3/5"], "correct": 1},
        {"q": "144 ning kvadrat ildizi?", "options": ["11","12","13","14"], "correct": 1},
        {"q": "3x=27, x=?", "options": ["7","8","9","10"], "correct": 2},
    ],
    6: [
        {"q": "-8+(-5)=?", "options": ["-13","-3","3","13"], "correct": 0},
        {"q": "2x+5=15, x=?", "options": ["4","5","6","7"], "correct": 1},
        {"q": "40% dan 350 necha?", "options": ["130","135","140","145"], "correct": 2},
        {"q": "(-3)×(-4)=?", "options": ["-12","-7","7","12"], "correct": 3},
        {"q": "15/20 ni qisqartiring", "options": ["3/4","1/2","4/5","2/3"], "correct": 0},
        {"q": "Radiusi 7 bo'lgan aylananing diametri?", "options": ["10","12","14","16"], "correct": 2},
        {"q": "5²-3²=?", "options": ["14","16","18","20"], "correct": 1},
        {"q": "3:4 = 15:x, x=?", "options": ["18","19","20","21"], "correct": 2},
        {"q": "-12÷4=?", "options": ["-4","-3","3","4"], "correct": 1},
        {"q": "2(x+3)=16, x=?", "options": ["4","5","6","7"], "correct": 1},
    ],
    7: [
        {"q": "3x-7=14, x=?", "options": ["5","6","7","8"], "correct": 2},
        {"q": "(x+2)(x-2)=?", "options": ["x²-4","x²+4","x²-2x","2x"], "correct": 0},
        {"q": "Ikki burchak yig'indisi 90°, biri 35°. Ikkinchisi?", "options": ["45","55","60","65"], "correct": 1},
        {"q": "2⁵=?", "options": ["16","32","64","10"], "correct": 1},
        {"q": "y=2x+1, x=3 bo'lsa y=?", "options": ["5","6","7","8"], "correct": 2},
        {"q": "√81=?", "options": ["7","8","9","10"], "correct": 2},
        {"q": "Uchburchak ichki burchaklar yig'indisi?", "options": ["90","180","270","360"], "correct": 1},
        {"q": "-3x=12, x=?", "options": ["-4","-3","3","4"], "correct": 0},
        {"q": "5(2x-1)=45, x=?", "options": ["4","5","6","7"], "correct": 1},
        {"q": "x/6=4/8, x=?", "options": ["2","3","4","5"], "correct": 1},
    ],
}

TEST_SUBJECTS = {
    "Matematika": MATH_QUESTIONS,
}
class DiagnosticTest(StatesGroup):
    question = State()

class RetryQuiz(StatesGroup):
    question = State()

QIYINLIK_HARFLAR = {"A": 0, "B": 1, "C": 2, "D": 3}
DIAGNOSTIC_CACHE = {}

def _load_diagnostic_sync(grade: int):
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]
    creds_json = os.getenv("GOOGLE_CREDS")
    creds_data = json.loads(creds_json)
    creds = Credentials.from_service_account_info(creds_data, scopes=scopes)
    client = gspread.authorize(creds)

    tests_sheet_url = os.getenv("TESTS_SHEET_URL")
    spreadsheet = client.open_by_url(tests_sheet_url)
    sheet_name = f"kimyo_{grade}sinf_diagnostik"
    sheet = spreadsheet.worksheet(sheet_name)
    rows = sheet.get_all_records()

    levels = {"oson": [], "orta": [], "murakkab": []}
    for row in rows:
        try:
            qiyinlik = str(row.get("Qiyinlik", "")).strip().lower()
            savol = str(row.get("Savol", "")).strip()
            options = [
                str(row.get("A", "")),
                str(row.get("B", "")),
                str(row.get("C", "")),
                str(row.get("D", "")),
            ]
            togri_harf = str(row.get("TogriJavob", "")).strip().upper()
            correct_idx = QIYINLIK_HARFLAR.get(togri_harf)
            izoh = str(row.get("Izoh", "")).strip()
            if correct_idx is None or qiyinlik not in levels:
                continue
            levels[qiyinlik].append({
                "q": savol, "options": options,
                "correct": correct_idx, "izoh": izoh
            })
        except Exception:
            logging.exception(f"Diagnostik qator xatosi: {row}")
            continue
    return levels

def get_diagnostic_questions(grade: int):
    if grade not in DIAGNOSTIC_CACHE:
        DIAGNOSTIC_CACHE[grade] = _load_diagnostic_sync(grade)
    levels = DIAGNOSTIC_CACHE[grade]
    selected = []
    for level in ("oson", "orta", "murakkab"):
        pool = levels.get(level, [])
        count = min(5, len(pool))
        selected.extend(random.sample(pool, count))
    random.shuffle(selected)
    return selected

async def start_diagnostic_test(message: types.Message, state: FSMContext, grade: int):
    try:
        questions = await asyncio.to_thread(get_diagnostic_questions, grade)
    except Exception:
        logging.exception("Diagnostik testni yuklashda xato:")
        await message.answer("⚠️ Diagnostik testni hozircha yuklab bo'lmadi, keyinroq urinib ko'ring.")
        return

    if not questions:
        await message.answer("⚠️ Bu sinf uchun diagnostik savollar hali tayyor emas.")
        return

    await state.update_data(diag_questions=questions, diag_index=0, diag_score=0, diag_wrong=[], grade=grade)
    await state.set_state(DiagnosticTest.question)
    await message.answer(
        f"🧪 {grade}-sinf Kimyo Diagnostik testi boshlanmoqda!\n\n"
        "📊 Bu sizning boshlang'ich bilim darajangizni aniqlash uchun 15 ta savoldan iborat. Boshladik!"
    )
    await send_diagnostic_question(message, state)

async def send_diagnostic_question(message: types.Message, state: FSMContext):
    data = await state.get_data()
    questions = data["diag_questions"]
    idx = data["diag_index"]
    q = questions[idx]

    kb = InlineKeyboardBuilder()
    letters = ["A", "B", "C", "D"]
    for i, opt in enumerate(q["options"]):
        kb.button(text=f"{letters[i]}) {opt}", callback_data=f"diag_ans_{i}")
    kb.adjust(1)

    await message.answer(
        f"❓ Savol {idx+1}/{len(questions)}:\n\n{q['q']}",
        reply_markup=kb.as_markup()
    )

@dp.callback_query(DiagnosticTest.question, F.data.startswith("diag_ans_"))
async def process_diagnostic_answer(callback: types.CallbackQuery, state: FSMContext):
    chosen = int(callback.data.replace("diag_ans_", ""))
    data = await state.get_data()
    questions = data["diag_questions"]
    idx = data["diag_index"]
    score = data["diag_score"]
    wrong = data["diag_wrong"]
    q = questions[idx]
    if chosen == q["correct"]:
        score += 1
        await callback.message.edit_text(f"✅ To'g'ri!\n\n{q['q']}")
    else:
        wrong.append(q)
        await callback.message.edit_text(f"❌ Xato deb belgilandi.\n\n{q['q']}")

    idx += 1
    await state.update_data(diag_index=idx, diag_score=score, diag_wrong=wrong)
    await callback.answer()

    if idx >= len(questions):
        await finish_diagnostic(callback.message, state)
    else:
        await send_diagnostic_question(callback.message, state)

async def finish_diagnostic(message: types.Message, state: FSMContext):
    data = await state.get_data()
    total = len(data["diag_questions"])
    score = data["diag_score"]
    wrong = data["diag_wrong"]
    grade = data.get("grade", "—")
    percent = round(score / total * 100)

    await save_test_result(message.chat.id, "Kimyo (Diagnostik)", grade, percent)
    await notify_admin_test_result(message.chat.id, "Kimyo (Diagnostik)", grade, score, total, percent)
    daraja = get_student_level(percent)

    await message.answer(
        f"🏁 Diagnostik test yakunlandi!\n\n"
        f"✅ To'g'ri javoblar: {score}/{total}\n"
        f"📊 Natija: {percent}%\n"
        f"🎖 Daraja: {daraja}\n\n"
        f"Bu sizning boshlang'ich darajangiz — darslar davomida uni yaxshilab boramiz!"
    )

    if wrong:
        await state.update_data(diag_wrong=wrong)
        kb = InlineKeyboardBuilder()
        kb.button(text="🆘 Yordam kerak", callback_data="show_diag_wrong")
        kb.button(text="💪 Mustaqil ishlayman", callback_data="hide_diag_wrong")
        kb.adjust(2)
        await message.answer(
            f"📝 Xatolar ustida ishlash\n\n"
            f"Siz {len(wrong)} ta savolda xato qildingiz.\n\n"
            f"Avval o'zingiz shu mavzularni qayta ko'rib, xatoni topishga harakat qiling! "
            f"Agar yordam kerak bo'lsa, pastdagi tugmani bosing.",
            reply_markup=kb.as_markup()
        )
    else:
        await state.clear()


@dp.callback_query(F.data == "show_diag_wrong")
async def show_diag_wrong(callback: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    wrong = data.get("diag_wrong", [])
    letters = ["A", "B", "C", "D"]
    lines = []
    for i, q in enumerate(wrong, 1):
        lines.append(
            f"{i}. {q['q']}\n"
            f"✅ To'g'ri javob: {letters[q['correct']]}) {q['options'][q['correct']]}\n"
            f"💡 {q['izoh']}"
        )
    text = "📖 Xato qilingan savollar va tushuntirishlar:\n\n" + "\n\n".join(lines)
    await callback.message.edit_text(text)
    await callback.answer()
    await state.clear()


MAX_RETRY_ATTEMPTS = 20


async def start_retry_loop(message, state: FSMContext, wrong_question_dicts: list):
    retry_queue = [{"q": q, "attempts": 0} for q in wrong_question_dicts]
    await state.update_data(retry_queue=retry_queue)
    await state.set_state(RetryQuiz.question)
    await message.answer(
        f"💪 Mustaqil mashq boshlandi!\n\n"
        f"Siz xato qilgan {len(retry_queue)} ta savolni ketma-ket qayta yechasiz. "
        f"Har bir savolni to'g'ri javoblaguningizcha shu savol qaytaveradi!"
    )
    await send_retry_question(message, state)


async def send_retry_question(message, state: FSMContext):
    data = await state.get_data()
    retry_queue = data.get("retry_queue", [])
    if not retry_queue:
        await message.answer(
            "🎉 Ajoyib! Siz xatolaringiz ustida ishlab, barcha savollarni to'g'ri javobladingiz!",
            reply_markup=get_main_menu()
        )
        await state.clear()
        return

    item = retry_queue[0]
    q = item["q"]
    kb = InlineKeyboardBuilder()
    for idx, option in enumerate(q["options"]):
        kb.button(text=option, callback_data=f"retry_ans_{idx}")
    max_len = max(len(opt) for opt in q["options"])
    kb.adjust(1 if max_len > 20 else 2)

    await message.answer(
        f"🔁 Qolgan: {len(retry_queue)} ta savol\n\n{q['q']}",
        reply_markup=kb.as_markup()
    )


@dp.callback_query(RetryQuiz.question, F.data.startswith("retry_ans_"))
async def process_retry_answer(callback: types.CallbackQuery, state: FSMContext):
    chosen = int(callback.data.replace("retry_ans_", ""))
    data = await state.get_data()
    retry_queue = data.get("retry_queue", [])
    item = retry_queue[0]
    q = item["q"]
    correct = q["correct"]

    if chosen == correct:
        await callback.message.edit_text(f"✅ To'g'ri!\n\n{q['q']}")
        retry_queue.pop(0)
    else:
        item["attempts"] += 1
        if item["attempts"] >= MAX_RETRY_ATTEMPTS:
            correct_text = q["options"][correct]
            izoh = q.get("izoh")
            extra = f"\n💡 {izoh}" if izoh else ""
            await callback.message.edit_text(
                f"❌ Xato.\n\n{q['q']}\n\n"
                f"{MAX_RETRY_ATTEMPTS} marta urinishdan so'ng — to'g'ri javob: {correct_text}{extra}\n\n"
                f"Bu mavzuni alohida qayta ko'rib chiqishingizni tavsiya qilamiz."
            )
            retry_queue.pop(0)
        else:
            await callback.message.edit_text(
                f"❌ Xato! Qayta urinib ko'ring.\n\n{q['q']}\n"
                f"(Urinish: {item['attempts']}/{MAX_RETRY_ATTEMPTS})"
            )

    await callback.answer()
    await state.update_data(retry_queue=retry_queue)
    await send_retry_question(callback.message, state)


@dp.callback_query(F.data == "hide_diag_wrong")
async def hide_diag_wrong(callback: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    wrong_dicts = data.get("diag_wrong", [])
    await callback.answer()
    await start_retry_loop(callback.message, state, wrong_dicts)


FAN_TO_BIRLASHMA = {
    "kimyo": "Tabiiy fanlar",
    "biologiya": "Tabiiy fanlar",
    "fizika": "Tabiiy fanlar",
    "jismoniy tarbiya": "Tabiiy fanlar",
    "matematika": "Aniq fanlar",
    "geometriya": "Aniq fanlar",
    "algebra": "Aniq fanlar",
    "informatika": "Aniq fanlar",
    "tarix": "Ijtimoiy fanlar",
    "huquq": "Ijtimoiy fanlar",
    "iqtisod": "Ijtimoiy fanlar",
    "chizmachilik": "Amaliy fanlar",
    "tasviriy san'at": "Amaliy fanlar",
    "rasm": "Amaliy fanlar",
    "mehnat ta'limi": "Amaliy fanlar",
    "mehnat talimi": "Amaliy fanlar",
    "ingliz tili": "Gumanitar fanlar",
    "rus tili": "Gumanitar fanlar",
    "nemis tili": "Gumanitar fanlar",
    "fransuz tili": "Gumanitar fanlar",
    "tojik tili": "Gumanitar fanlar",
    "koreys tili": "Gumanitar fanlar",
    "ona tili": "Gumanitar fanlar",
    "adabiyot": "Gumanitar fanlar",
}


def get_birlashma_for_fan(fan: str):
    if not fan:
        return None
    fan_lower = fan.strip().lower()
    for key, birlashma in FAN_TO_BIRLASHMA.items():
        if key in fan_lower:
            return birlashma
    return None


def _get_role_sync(user_id: int):
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]
    creds_json = os.getenv("GOOGLE_CREDS")
    if creds_json:
        creds_data = json.loads(creds_json)
        creds = Credentials.from_service_account_info(creds_data, scopes=scopes)
    else:
        creds = Credentials.from_service_account_file("credentials.json", scopes=scopes)
    client = gspread.authorize(creds)

    sheet_url = os.getenv("GOOGLE_SHEET_URL")
    spreadsheet_id = os.getenv("SPREADSHEET_ID")
    if sheet_url:
        spreadsheet = client.open_by_url(sheet_url)
    elif spreadsheet_id:
        spreadsheet = client.open_by_key(spreadsheet_id)
    else:
        spreadsheet = client.open_by_url(
            "https://docs.google.com/spreadsheets/d/1aXoL-TeP0Oh62u1kfgPyzyRsNjOdqGkovJmFutYlUn0/edit"
        )

    try:
        sheet = spreadsheet.worksheet("Rollar")
    except Exception:
        sheet = spreadsheet.add_worksheet(title="Rollar", rows=200, cols=7)
        sheet.append_row(["Telegram ID", "Rol", "Maktab", "Sinf", "Metod Birlashma", "Fan", "Ism"])
        return None

    for row in sheet.get_all_records():
        try:
            if int(row.get("Telegram ID")) == user_id:
                sinf = str(row.get("Sinf", "")).strip()
                metod_birlashma = str(row.get("Metod Birlashma", "")).strip()
                fan = str(row.get("Fan", "")).strip()
                if not metod_birlashma and fan:
                    metod_birlashma = get_birlashma_for_fan(fan) or ""
                return {
                    "role": str(row.get("Rol", "")).strip().lower(),
                    "maktab": str(row.get("Maktab", "")).strip(),
                    "sinf": sinf if sinf else None,
                    "metod_birlashma": metod_birlashma if metod_birlashma else None,
                    "fan": fan if fan else None
                }
        except (ValueError, TypeError):
            continue
    return None


async def get_user_role(user_id: int):
    try:
        return await asyncio.to_thread(_get_role_sync, user_id)
    except Exception:
        logging.exception("Rolni tekshirishda xato:")
        return None


ROLE_LABELS = {
    "direktor": "🏫 Direktor paneli",
    "oibdo": "📘 O'IBDO' paneli (o'quv ishlari)",
    "otibdo": "📗 O'TIBDO' paneli (o'quv-tarbiyaviy ishlar)",
    "texnik": "🔧 Texnik xodimlar bo'limi paneli",
    "boshlangich_oqituvchi": "🧒 Boshlang'ich sinf o'qituvchisi paneli",
    "yuqori_sinf_oqituvchi": "🎓 Yuqori sinf o'qituvchisi paneli",
    "birlashma_rahbari": "🧩 Metod birlashma rahbari paneli",
    "sinf_rahbari": "👨‍👩‍👧 Sinf rahbari paneli",
}

REQUESTABLE_ROLES = {
    "direktor": "🏫 Direktor",
    "oibdo": "📘 O'IBDO'",
    "otibdo": "📗 O'TIBDO'",
    "texnik": "🔧 Texnik xodim",
    "boshlangich_oqituvchi": "🧒 Boshlang'ich sinf o'qituvchisi",
    "yuqori_sinf_oqituvchi": "🎓 Yuqori sinf o'qituvchisi",
    "birlashma_rahbari": "🧩 Metod birlashma rahbari",
    "aa_ustoz": "🎓 AA Ustozi (Angren Akademiyasi xodimi)",
    "sinf_rahbari": "👨‍👩‍👧 Sinf rahbari",
}

# Bu rollar butun maktab statistikasini ko'radi
SCHOOL_WIDE_ROLES = {"direktor", "oibdo", "otibdo", "texnik"}
# Bu rollar faqat o'z sinfi statistikasini ko'radi (bitta sinfga hamma fanni o'qitadi / sinf rahbari)
CLASS_SCOPED_ROLES = {"boshlangich_oqituvchi", "sinf_rahbari"}
# Bu rol faqat o'zi o'qitadigan FAN bo'yicha, butun maktab kesimida ko'radi
SUBJECT_SCOPED_ROLES = {"yuqori_sinf_oqituvchi"}
# Bu rol o'z metod birlashmasiga tegishli barcha fanlar bo'yicha ko'radi
BIRLASHMA_LEADER_ROLES = {"birlashma_rahbari"}

BOSHLANGICH_TALIM_NOMLARI = {"boshlang'ich ta'lim", "boshlangich talim", "boshlang'ich talim"}


def _get_birlashma_fanlar_sync(maktab: str, metod_birlashma: str):
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]
    creds_json = os.getenv("GOOGLE_CREDS")
    if creds_json:
        creds_data = json.loads(creds_json)
        creds = Credentials.from_service_account_info(creds_data, scopes=scopes)
    else:
        creds = Credentials.from_service_account_file("credentials.json", scopes=scopes)
    client = gspread.authorize(creds)

    sheet_url = os.getenv("GOOGLE_SHEET_URL")
    spreadsheet_id = os.getenv("SPREADSHEET_ID")
    if sheet_url:
        spreadsheet = client.open_by_url(sheet_url)
    elif spreadsheet_id:
        spreadsheet = client.open_by_key(spreadsheet_id)
    else:
        spreadsheet = client.open_by_url(
            "https://docs.google.com/spreadsheets/d/1aXoL-TeP0Oh62u1kfgPyzyRsNjOdqGkovJmFutYlUn0/edit"
        )

    try:
        sheet = spreadsheet.worksheet("Rollar")
    except Exception:
        return []

    birlashma_lower = metod_birlashma.strip().lower()
    fanlar = set()
    for row in sheet.get_all_records():
        row_maktab = str(row.get("Maktab", "")).strip()
        row_fan = str(row.get("Fan", "")).strip()
        row_birlashma = str(row.get("Metod Birlashma", "")).strip()
        if not row_birlashma and row_fan:
            row_birlashma = get_birlashma_for_fan(row_fan) or ""
        if row_maktab == maktab and row_birlashma.strip().lower() == birlashma_lower and row_fan:
            fanlar.add(row_fan)
    return list(fanlar)


async def get_birlashma_fanlar(maktab: str, metod_birlashma: str):
    try:
        return await asyncio.to_thread(_get_birlashma_fanlar_sync, maktab, metod_birlashma)
    except Exception:
        logging.exception("Birlashma fanlarini olishda xato:")
        return []


def _get_scope_stats_sync(maktab: str, sinf: str = None, fan=None, grade_range=None):
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]
    creds_json = os.getenv("GOOGLE_CREDS")
    if creds_json:
        creds_data = json.loads(creds_json)
        creds = Credentials.from_service_account_info(creds_data, scopes=scopes)
    else:
        creds = Credentials.from_service_account_file("credentials.json", scopes=scopes)
    client = gspread.authorize(creds)

    sheet_url = os.getenv("GOOGLE_SHEET_URL")
    spreadsheet_id = os.getenv("SPREADSHEET_ID")
    if sheet_url:
        spreadsheet = client.open_by_url(sheet_url)
    elif spreadsheet_id:
        spreadsheet = client.open_by_key(spreadsheet_id)
    else:
        spreadsheet = client.open_by_url(
            "https://docs.google.com/spreadsheets/d/1aXoL-TeP0Oh62u1kfgPyzyRsNjOdqGkovJmFutYlUn0/edit"
        )

    try:
        sheet = spreadsheet.worksheet("Test Natijalari")
    except Exception:
        return None

    rows = [r for r in sheet.get_all_records() if str(r.get("Maktab", "")).strip() == maktab]

    if sinf:
        rows = [r for r in rows if str(r.get("Sinf", "")).strip() == str(sinf).strip()]

    if fan:
        fan_list = [fan] if isinstance(fan, str) else list(fan)
        fan_list_lower = [f.strip().lower() for f in fan_list if f]
        rows = [
            r for r in rows
            if any(fl in str(r.get("Fan", "")).strip().lower() for fl in fan_list_lower)
        ]

    if grade_range:
        min_g, max_g = grade_range
        filtered = []
        for r in rows:
            try:
                g = int(str(r.get("Sinf", "")).strip())
                if min_g <= g <= max_g:
                    filtered.append(r)
            except (ValueError, TypeError):
                continue
        rows = filtered

    if not rows:
        return {"jami": 0}

    counts = {"🏆 Mutlaq a'lochi": 0, "🥇 A'lochi o'quvchi": 0, "🥈 Yaxshi bilimli": 0,
              "🥉 O'rtacha bilimli": 0, "⚠️ Past o'zlashtiruvchi": 0}
    total_percent = 0
    for r in rows:
        daraja = str(r.get("Daraja", "")).strip()
        if daraja in counts:
            counts[daraja] += 1
        try:
            total_percent += float(r.get("Natija (%)", 0))
        except (ValueError, TypeError):
            pass

    return {
        "jami": len(rows),
        "mutlaq_alochi": counts["🏆 Mutlaq a'lochi"],
        "alochi": counts["🥇 A'lochi o'quvchi"],
        "yaxshi": counts["🥈 Yaxshi bilimli"],
        "ortacha": counts["🥉 O'rtacha bilimli"],
        "past": counts["⚠️ Past o'zlashtiruvchi"],
        "ortacha_foiz": round(total_percent / len(rows), 1)
    }


async def get_scope_stats(maktab: str, sinf: str = None, fan=None, grade_range=None):
    try:
        return await asyncio.to_thread(_get_scope_stats_sync, maktab, sinf, fan, grade_range)
    except Exception:
        logging.exception("Statistikani olishda xato:")
        return None


def _add_role_sync(user_id: int, full_name: str, role: str, maktab: str, sinf: str = ""):
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]
    creds_json = os.getenv("GOOGLE_CREDS")
    if creds_json:
        creds_data = json.loads(creds_json)
        creds = Credentials.from_service_account_info(creds_data, scopes=scopes)
    else:
        creds = Credentials.from_service_account_file("credentials.json", scopes=scopes)
    client = gspread.authorize(creds)

    sheet_url = os.getenv("GOOGLE_SHEET_URL")
    spreadsheet_id = os.getenv("SPREADSHEET_ID")
    if sheet_url:
        spreadsheet = client.open_by_url(sheet_url)
    elif spreadsheet_id:
        spreadsheet = client.open_by_key(spreadsheet_id)
    else:
        spreadsheet = client.open_by_url(
            "https://docs.google.com/spreadsheets/d/1aXoL-TeP0Oh62u1kfgPyzyRsNjOdqGkovJmFutYlUn0/edit"
        )

    try:
        sheet = spreadsheet.worksheet("Rollar")
    except Exception:
        sheet = spreadsheet.add_worksheet(title="Rollar", rows=200, cols=7)
        sheet.append_row(["Telegram ID", "Rol", "Maktab", "Sinf", "Metod Birlashma", "Fan", "Ism"])

    sheet.append_row([user_id, role, maktab, sinf, "", "", full_name])


async def add_role(user_id: int, full_name: str, role: str, maktab: str, sinf: str = ""):
    try:
        await asyncio.to_thread(_add_role_sync, user_id, full_name, role, maktab, sinf)
        return True
    except Exception:
        logging.exception("Rol qo'shishda xato:")
        return False


def _find_director_sync(maktab: str):
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]
    creds_json = os.getenv("GOOGLE_CREDS")
    if creds_json:
        creds_data = json.loads(creds_json)
        creds = Credentials.from_service_account_info(creds_data, scopes=scopes)
    else:
        creds = Credentials.from_service_account_file("credentials.json", scopes=scopes)
    client = gspread.authorize(creds)

    sheet_url = os.getenv("GOOGLE_SHEET_URL")
    spreadsheet_id = os.getenv("SPREADSHEET_ID")
    if sheet_url:
        spreadsheet = client.open_by_url(sheet_url)
    elif spreadsheet_id:
        spreadsheet = client.open_by_key(spreadsheet_id)
    else:
        spreadsheet = client.open_by_url(
            "https://docs.google.com/spreadsheets/d/1aXoL-TeP0Oh62u1kfgPyzyRsNjOdqGkovJmFutYlUn0/edit"
        )

    try:
        sheet = spreadsheet.worksheet("Rollar")
    except Exception:
        return None

    for row in sheet.get_all_records():
        if str(row.get("Maktab", "")).strip() == maktab and str(row.get("Rol", "")).strip().lower() == "direktor":
            try:
                return int(row.get("Telegram ID"))
            except (ValueError, TypeError):
                continue
    return None


async def find_director(maktab: str):
    try:
        return await asyncio.to_thread(_find_director_sync, maktab)
    except Exception:
        logging.exception("Direktorni qidirishda xato:")
        return None


def get_main_menu(has_panel: bool = False):
    kb = ReplyKeyboardBuilder()
    kb.button(text="📝 Ro'yxatdan o'tish")
    kb.button(text="📈 Bilim darajasini tekshirish")
    kb.button(text="🚪 Davomat (Keldim/Ketdim)")
    kb.adjust(1, 2)
    kb.button(text="👤 Shaxsiy kabinet")
    kb.button(text="🔑 Boshqaruv huquqini so'rash")
    if has_panel:
        kb.button(text="🏫 Maktab paneli")
    return kb.as_markup(resize_keyboard=True)


@dp.message(F.text == "/start")
async def cmd_start(message: types.Message, state: FSMContext):
    await state.clear()
    if not await check_subscription(message.from_user.id):
        await message.answer(
            "📍 Botdan foydalanish uchun avval bizning rasmiy kanalimizga a'zo bo'ling:",
            reply_markup=get_subscribe_keyboard()
        )
        return
    role_info = await get_user_role(message.from_user.id)
    has_panel = bool(role_info and role_info.get("role") in ROLE_LABELS)
    await message.answer(
        "✨ Angren Akademiyasi rasmiy botiga xush kelibsiz!\n\n"
        "Kelajak akademiyasida o'z bilimingizni va farzandingiz kamolotini nazorat qiling.",
        reply_markup=get_main_menu(has_panel)
    )


@dp.message(F.text == "🏫 Maktab paneli")
async def school_panel(message: types.Message):
    role_info = await get_user_role(message.from_user.id)
    role = role_info.get("role") if role_info else None
    if not role_info or role not in ROLE_LABELS:
        await message.answer("❌ Sizda bu bo'limga kirish huquqi yo'q.")
        return

    maktab = role_info["maktab"]
    sinf = role_info.get("sinf")
    fan = role_info.get("fan")
    metod_birlashma = role_info.get("metod_birlashma")

    filter_sinf = None
    filter_fan = None
    filter_grade_range = None
    birlashma_fanlar = []

    if role in CLASS_SCOPED_ROLES:
        if not sinf:
            await message.answer("⚠️ Sizning \"Sinf\" ma'lumotingiz Rollar jadvalida kiritilmagan. Admin bilan bog'laning.")
            return
        filter_sinf = sinf
    elif role in SUBJECT_SCOPED_ROLES:
        if not fan:
            await message.answer("⚠️ Sizning \"Fan\" ma'lumotingiz Rollar jadvalida kiritilmagan. Admin bilan bog'laning.")
            return
        filter_fan = fan
    elif role in BIRLASHMA_LEADER_ROLES:
        if not metod_birlashma:
            await message.answer("⚠️ Sizning \"Metod Birlashma\" ma'lumotingiz Rollar jadvalida kiritilmagan. Admin bilan bog'laning.")
            return
        if metod_birlashma.strip().lower() in BOSHLANGICH_TALIM_NOMLARI:
            filter_grade_range = (1, 4)
        else:
            birlashma_fanlar = await get_birlashma_fanlar(maktab, metod_birlashma)
            filter_fan = birlashma_fanlar if birlashma_fanlar else None

    stats = await get_scope_stats(maktab, sinf=filter_sinf, fan=filter_fan, grade_range=filter_grade_range)

    if not stats:
        await message.answer("⚠️ Statistikani yuklab bo'lmadi, keyinroq urinib ko'ring.")
        return

    header_lines = [ROLE_LABELS[role], "", f"🏫 {maktab}"]
    if role in SUBJECT_SCOPED_ROLES:
        header_lines.append(f"📚 Fan: {fan}")
        if metod_birlashma:
            header_lines.append(f"🧬 Metod birlashma: {metod_birlashma}")
        if sinf:
            header_lines.append(f"👨‍👩‍👧 Sinf rahbarligi: {sinf}-sinf")
    elif role in CLASS_SCOPED_ROLES:
        header_lines.append(f"🏷 Sinf: {sinf}-sinf")
        if metod_birlashma:
            header_lines.append(f"🧬 Metod birlashma: {metod_birlashma}")
    elif role in BIRLASHMA_LEADER_ROLES:
        header_lines.append(f"🧩 Metod birlashma: {metod_birlashma}")
        if filter_grade_range:
            header_lines.append("🏷 Qamrov: 1–4-sinflar")
        elif birlashma_fanlar:
            header_lines.append(f"📚 Fanlar: {', '.join(birlashma_fanlar)}")
        else:
            header_lines.append("📚 Fanlar: hali biriktirilmagan")

    header_text = "\n".join(header_lines)

    if stats["jami"] == 0:
        await message.answer(f"{header_text}\n\nHozircha test natijalari mavjud emas.")
        return

    await message.answer(
        f"{header_text}\n\n"
        f"📊 Jami test natijalari: {stats['jami']} ta\n\n"
        f"🏆 Mutlaq a'lochi: {stats['mutlaq_alochi']}\n"
        f"🥇 A'lochi o'quvchi: {stats['alochi']}\n"
        f"🥈 Yaxshi bilimli: {stats['yaxshi']}\n"
        f"🥉 O'rtacha bilimli: {stats['ortacha']}\n"
        f"⚠️ Past o'zlashtiruvchi: {stats['past']}\n\n"
        f"📈 O'rtacha ko'rsatkich: {stats['ortacha_foiz']}%"
    )


@dp.message(F.text == "🔑 Boshqaruv huquqini so'rash")
async def request_role_start(message: types.Message, state: FSMContext):
    kb = InlineKeyboardBuilder()
    for idx, school in enumerate(ANGREN_SCHOOLS):
        kb.button(text=school, callback_data=f"rrmaktab_{idx}")
    kb.adjust(4)
    await message.answer("🏫 Qaysi maktab uchun huquq so'rayapsiz?", reply_markup=kb.as_markup())
    await state.set_state(RoleRequest.maktab)


@dp.callback_query(RoleRequest.maktab, F.data.startswith("rrmaktab_"))
async def request_role_school(callback: types.CallbackQuery, state: FSMContext):
    idx = int(callback.data.replace("rrmaktab_", ""))
    maktab = ANGREN_SCHOOLS[idx]
    await state.update_data(maktab=maktab)

    kb = InlineKeyboardBuilder()
    for key, label in REQUESTABLE_ROLES.items():
        kb.button(text=label, callback_data=f"rrrole_{key}")
    kb.adjust(1)
    await callback.message.edit_text(f"🏫 Tanlandi: {maktab}\n\nQaysi lavozim uchun so'rayapsiz?", reply_markup=kb.as_markup())
    await state.set_state(RoleRequest.role)
    await callback.answer()


@dp.callback_query(RoleRequest.role, F.data.startswith("rrrole_"))
async def request_role_final(callback: types.CallbackQuery, state: FSMContext):
    role_key = callback.data.replace("rrrole_", "")
    await state.update_data(role_key=role_key)
    await callback.answer()

    if role_key in CLASS_SCOPED_ROLES:
        await callback.message.edit_text("👨‍👩‍👧 Qaysi sinfga mas'ulsiz? (masalan: 7-A)")
        await state.set_state(RoleRequest.sinf)
        return

    await finalize_role_request(callback.message, callback.from_user, state, sinf="")


@dp.message(RoleRequest.sinf)
async def request_role_sinf(message: types.Message, state: FSMContext):
    await finalize_role_request(message, message.from_user, state, sinf=message.text.strip())


async def finalize_role_request(message: types.Message, user, state: FSMContext, sinf: str):
    data = await state.get_data()
    maktab = data.get("maktab")
    role_key = data.get("role_key")
    role_label = REQUESTABLE_ROLES.get(role_key, role_key)
    full_name = user.full_name or "Noma'lum"

    sinf_line = f"\n🏷 Sinf: {sinf}" if sinf else ""
    await message.answer(
        f"✅ So'rovingiz yuborildi!\n\n🏫 {maktab}\n👔 {role_label}{sinf_line}\n\nTasdiqlanishini kuting."
    )
    await state.clear()

    admin_id = os.getenv("ADMIN_ID")

    if role_key in ("direktor", "aa_ustoz"):
        approver_id = int(admin_id) if admin_id else None
    else:
        approver_id = await find_director(maktab)
        if not approver_id and admin_id:
            approver_id = int(admin_id)

    if approver_id:
        kb = InlineKeyboardBuilder()
        kb.button(text="✅ Tasdiqlash", callback_data=f"rrapprove_{user.id}_{role_key}_{sinf or '-'}")
        kb.button(text="❌ Rad etish", callback_data=f"rrreject_{user.id}")
        kb.adjust(2)
        try:
            await bot.send_message(
                approver_id,
                f"🔑 YANGI HUQUQ SO'ROVI\n\n"
                f"👤 {full_name} (@{user.username or '—'})\n"
                f"🆔 ID: {user.id}\n"
                f"🏫 Maktab: {maktab}\n"
                f"👔 Lavozim: {role_label}{sinf_line}",
                reply_markup=kb.as_markup()
            )
        except Exception:
            logging.exception("Tasdiqlovchiga so'rov yuborishda xato:")


@dp.callback_query(F.data.startswith("rrapprove_"))
async def approve_role_request(callback: types.CallbackQuery):
    admin_id = os.getenv("ADMIN_ID", "")
    caller_id = str(callback.from_user.id)

    _, user_id_str, role_key, sinf_raw = callback.data.split("_", 3)
    user_id = int(user_id_str)
    sinf = "" if sinf_raw == "-" else sinf_raw

    text = callback.message.text
    maktab = "—"
    for line in text.split("\n"):
        if line.startswith("🏫 Maktab:"):
            maktab = line.replace("🏫 Maktab:", "").strip()
            break

    # Ruxsat tekshiruvi: direktor/AA ustozi so'rovini faqat Rahbar tasdiqlaydi;
    # boshqa maktab xodimlari so'rovini o'sha maktabning direktori (yoki Rahbar) tasdiqlaydi
    if role_key in ("direktor", "aa_ustoz"):
        if caller_id != admin_id:
            await callback.answer("❌ Sizda bu huquq yo'q.", show_alert=True)
            return
    else:
        director_id = await find_director(maktab)
        allowed_ids = {admin_id}
        if director_id:
            allowed_ids.add(str(director_id))
        if caller_id not in allowed_ids:
            await callback.answer("❌ Sizda bu huquq yo'q.", show_alert=True)
            return

    try:
        chat = await bot.get_chat(user_id)
        full_name = chat.full_name or "Noma'lum"
    except Exception:
        full_name = "Noma'lum"

    success = await add_role(user_id, full_name, role_key, maktab, sinf)
    if success:
        await callback.message.edit_text(callback.message.text + "\n\n✅ TASDIQLANDI")
        try:
            await bot.send_message(user_id, f"🎉 Tabriklaymiz! Sizga \"{REQUESTABLE_ROLES.get(role_key, role_key)}\" huquqi berildi.\n\nEndi \"🏫 Maktab paneli\" tugmasi orqali kira olasiz (avval /start bosing).")
        except Exception:
            pass
    else:
        await callback.message.edit_text(callback.message.text + "\n\n⚠️ Xatolik yuz berdi, qayta urinib ko'ring.")
    await callback.answer()


@dp.callback_query(F.data.startswith("rrreject_"))
async def reject_role_request(callback: types.CallbackQuery):
    admin_id = os.getenv("ADMIN_ID", "")
    if str(callback.from_user.id) != admin_id:
        # direktor ham rad eta olishi mumkin — u faqat o'ziga yuborilgan so'rovlarga javob beradi,
        # shuning uchun bu yerda faqat Rahbarga tegishli qat'iy tekshiruv shart emas
        pass
    user_id = int(callback.data.replace("rrreject_", ""))
    await callback.message.edit_text(callback.message.text + "\n\n❌ RAD ETILDI")
    try:
        await bot.send_message(user_id, "❌ Afsuski, so'rovingiz rad etildi.")
    except Exception:
        pass
    await callback.answer()


@dp.message(F.text == "/help")
async def cmd_help(message: types.Message):
    await message.answer("Sizga qanday yordam bera olaman?")


@dp.message(F.text == "/excel")
async def send_excel(message: types.Message):
    if os.path.exists(EXCEL_FILE):
        excel_doc = FSInputFile(EXCEL_FILE)
        await message.answer_document(document=excel_doc, caption="📊 Angren Akademiya o'quvchilar ro'yxati (Excel)")
    else:
        await message.answer("❌ Hozircha ro'yxat bo'sh! Hech kim ro'yxatdan o'tmadi.")


@dp.message(F.text == "📈 Bilim darajasini tekshirish")
async def check_knowledge(message: types.Message, state: FSMContext):
    kb = InlineKeyboardBuilder()
    for subject in TEST_SUBJECTS:
        kb.button(text=subject, callback_data=f"test_subj_{subject}")
    kb.button(text="Kimyo", callback_data="test_subj_Kimyo")
    kb.adjust(1)
    await message.answer("📚 Qaysi fandan test topshirmoqchisiz?", reply_markup=kb.as_markup())
    await state.set_state(TestQuiz.subject)


@dp.callback_query(TestQuiz.subject, F.data.startswith("test_subj_"))
async def choose_grade(callback: types.CallbackQuery, state: FSMContext):
    subject = callback.data.replace("test_subj_", "")
    await state.update_data(subject=subject, score=0, q_index=0)

    kb = InlineKeyboardBuilder()
    grade_range = range(7, 12) if subject == "Kimyo" else range(1, 8)
    for grade in grade_range:
        kb.button(text=f"{grade}-sinf", callback_data=f"test_grade_{grade}")
    kb.adjust(4, 3)
    await callback.message.edit_text("🎓 Sinfingizni tanlang:", reply_markup=kb.as_markup())
    await state.set_state(TestQuiz.grade)
    await callback.answer()


@dp.callback_query(TestQuiz.grade, F.data.startswith("test_grade_"))
async def start_questions(callback: types.CallbackQuery, state: FSMContext):
    grade = int(callback.data.replace("test_grade_", ""))
    data = await state.get_data()
    subject = data["subject"]
    if subject == "Kimyo":
        questions = await asyncio.to_thread(get_diagnostic_questions, grade)
        if not questions:
            await callback.message.edit_text("⚠️ Bu sinf uchun Kimyo savollari hali tayyor emas.")
            await callback.answer()
            await state.clear()
            return
    else:
        questions = TEST_SUBJECTS[subject][grade]

    await state.update_data(grade=grade, questions=questions, q_index=0, score=0, wrong_questions=[])
    await state.set_state(TestQuiz.question)
    await callback.answer()
    label = "Diagnostik testi" if subject == "Kimyo" else "testi"
    await callback.message.answer(f"🧪 {grade}-sinf {subject} {label} boshlanmoqda! Omad tilaymiz 🍀")
    await send_question(callback.message, state)


async def send_question(message, state: FSMContext):
    data = await state.get_data()
    q_index = data["q_index"]
    questions = data["questions"]

    if q_index >= len(questions):
        await finish_test(message, state)
        return

    question = questions[q_index]
    kb = InlineKeyboardBuilder()
    for idx, option in enumerate(question["options"]):
        kb.button(text=option, callback_data=f"test_ans_{idx}")
    max_len = max(len(opt) for opt in question["options"])
    kb.adjust(1 if max_len > 20 else 2)

    await message.answer(
        f"❓ Savol {q_index + 1}/{len(questions)}:\n\n{question['q']}",
        reply_markup=kb.as_markup()
    )


@dp.callback_query(TestQuiz.question, F.data.startswith("test_ans_"))
async def process_answer(callback: types.CallbackQuery, state: FSMContext):
    chosen = int(callback.data.replace("test_ans_", ""))
    data = await state.get_data()
    q_index = data["q_index"]
    questions = data["questions"]
    score = data["score"]
    wrong_questions = data.get("wrong_questions", [])

    question = questions[q_index]
    correct = question["correct"]
    if chosen == correct:
        score += 1
        mark = "✅ To'g'ri!"
    else:
        correct_text = question["options"][correct]
        mark = f"❌ Xato! Siz tanladingiz: {question['options'][chosen]}\nTo'g'ri javob: {correct_text}"
        wrong_questions.append(q_index)

    try:
        await callback.message.edit_text(
            f"❓ Savol {q_index + 1}/{len(questions)}:\n\n{question['q']}\n\n{mark}"
        )
    except Exception:
        pass

    await callback.answer()
    await state.update_data(score=score, q_index=q_index + 1, wrong_questions=wrong_questions)
    await send_question(callback.message, state)


async def finish_test(message, state: FSMContext):
    data = await state.get_data()
    score = data["score"]
    questions = data["questions"]
    subject = data["subject"]
    grade = data["grade"]
    total = len(questions)
    percent = int((score / total) * 100)

    await save_test_result(message.chat.id, subject, grade, percent)
    await notify_admin_test_result(message.chat.id, subject, grade, score, total, percent)

    filename, title = get_diploma_tier(percent)
    daraja = get_student_level(percent)

    result_text = (
        f"🎉 Test yakunlandi!\n\n"
        f"📊 Natija: {score}/{total} ({percent}%)\n"
        f"🎖 Daraja: {daraja}\n"
    )

    if title:
        result_text += f"\n🏅 Siz \"{title}\" nominatsiyasiga munosib bo'ldingiz!"
    else:
        result_text += "\n💪 Yana urinib ko'ring, natijangizni yaxshilay olasiz!"

    await message.answer(result_text, reply_markup=get_main_menu())

    if filename:
        user_data = await state.get_data()
        found_name = await asyncio.to_thread(find_full_name_by_id, message.chat.id)
        full_name = found_name or message.chat.first_name or "O'quvchi"
        try:
            diploma_path = generate_diploma(full_name, f"{subject} ({grade}-sinf)", percent, message.chat.id)
            if diploma_path:
                await message.answer_photo(
                    photo=FSInputFile(diploma_path),
                    caption=f"🏅 \"{title}\" — Maqtov yorlig'ingiz tayyor!"
                )
                os.remove(diploma_path)
        except Exception:
            logging.exception("Diplom generatsiyasida xato:")

    wrong_questions = data.get("wrong_questions", [])
    if wrong_questions:
        nums = ", ".join(str(i + 1) for i in wrong_questions)
        kb = InlineKeyboardBuilder()
        kb.button(text="🆘 Yordam kerak", callback_data="show_wrong_answers")
        kb.button(text="💪 Mustaqil ishlayman", callback_data="hide_wrong_answers")
        kb.adjust(2)
        await message.answer(
            f"📝 Xatolar ustida ishlash\n\n"
            f"Siz {nums}-savollarda xato qildingiz.\n\n"
            f"Avval o'zingiz shu savollarni qayta ko'rib, xatoni topishga harakat qiling! "
            f"Agar yordam kerak bo'lsa, pastdagi tugmani bosing.",
            reply_markup=kb.as_markup()
        )
    else:
        await state.clear()

@dp.callback_query(F.data == "show_wrong_answers")
async def show_wrong_answers(callback: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    questions = data.get("questions", [])
    wrong_questions = data.get("wrong_questions", [])
    lines = []
    for i in wrong_questions:
        q = questions[i]
        correct_text = q["options"][q["correct"]]
        lines.append(f"{i + 1}. {q['q']}\n✅ To'g'ri javob: {correct_text}")
    text = "📖 Xato qilingan savollar va to'g'ri javoblar:\n\n" + "\n\n".join(lines)
    await callback.message.edit_text(text)
    await callback.answer()
    await state.clear()


@dp.callback_query(F.data == "hide_wrong_answers")
async def hide_wrong_answers(callback: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    questions = data.get("questions", [])
    wrong_questions = data.get("wrong_questions", [])
    wrong_dicts = [questions[i] for i in wrong_questions]
    await callback.answer()
    await start_retry_loop(callback.message, state, wrong_dicts)


@dp.message(F.text == "🚪 Davomat (Keldim/Ketdim)")
async def attendance_menu(message: types.Message):
    kb = InlineKeyboardBuilder()
    kb.button(text="🔔 Keldim", callback_data="attendance_in")
    kb.button(text="🔕 Ketdim", callback_data="attendance_out")
    kb.button(text="💰 Oylik to'lov holati", callback_data="attendance_pay")
    kb.button(text="🚀 Uzoq muddatli imtiyozlar", callback_data="attendance_promo")
    kb.button(text="👤 Shaxsiy kabinet", callback_data="profile")
    kb.adjust(2, 2, 1)
    await message.answer(
        "🚪 Angren Akademiyasi — Davomat va Shaxsiy Balans\n\nKerakli tugmani bosing:",
        reply_markup=kb.as_markup()
    )


@dp.callback_query(F.data.startswith("attendance_"))
async def process_attendance(callback: types.CallbackQuery):
    action = callback.data.split("_")[1]
    current_time = datetime.now().strftime("%H:%M")

    if action == "in":
        await callback.message.answer(
            f"🔔 Angren Akademiya Xabarnomasi\n\nFarzandingiz soat {current_time} da markazimizga eson-omon yetib keldi. 🔬"
        )
    elif action == "out":
        await callback.message.answer(
            f"🔕 Angren Akademiya Xabarnomasi\n\nFarzandingiz soat {current_time} da darsdan chiqdi. Oq yo'l! ☀️"
        )
    elif action == "pay":
        pay_kb = InlineKeyboardBuilder()
        pay_kb.button(text="💳 Plastik (Click/Payme)", callback_data="pay_via_card")
        pay_kb.button(text="💵 Naqd pul (Qo'lda)", callback_data="pay_via_cash")
        pay_kb.adjust(1)
        await callback.message.answer("💰 To'lov usulini tanlang:", reply_markup=pay_kb.as_markup())
    elif action == "promo":
        await callback.message.answer(
            "🚀 \"Angren Akademiya\" Premium Imtiyozlar:\n\n"
            "🥈 3 Oylik: 10% chegirma + sovg'a daftar 🎁\n"
            "🥇 6 Oylik: 15% chegirma + futbolka va kepka 👕\n"
            "👑 1 Yillik: 20% chegirma + darsliklar bepul 📚"
        )
    await callback.answer()


@dp.callback_query(F.data.startswith("pay_via_"))
async def method_payment(callback: types.CallbackQuery):
    method = callback.data.split("_")[2]
    karta_raqam = os.getenv("KARTA_RAQAMI", "8600 0000 0000 0000")
    karta_egasi = os.getenv("KARTA_EGASI", "Angren Akademiya Mas'ul Xodimi")

    if method == "card":
        await callback.message.answer(
            f"💳 Karta raqami: {karta_raqam}\nEga: {karta_egasi}\n\nChekni adminga yuboring."
        )
    else:
        await callback.message.answer("💵 To'lovni administratorga topshiring. Rahmat!")
    await callback.answer()


@dp.message(F.text == "📝 Ro'yxatdan o'tish")
async def start_registration(message: types.Message, state: FSMContext):
    if not await check_subscription(message.from_user.id):
        await message.answer(
            "📍 Ro'yxatdan o'tish uchun avval kanalimizga a'zo bo'ling:",
            reply_markup=get_subscribe_keyboard()
        )
        return
    await message.answer("Ism va familiyangizni kiriting:")
    await state.set_state(Registration.name)


@dp.message(Registration.name)
async def process_name(message: types.Message, state: FSMContext):
    await state.update_data(name=message.text)
    await message.answer("📞 O'quvchining telefon raqamini kiriting:")
    await state.set_state(Registration.phone)


@dp.message(Registration.phone)
async def process_phone(message: types.Message, state: FSMContext):
    await state.update_data(phone=message.text)
    await message.answer("👨‍👩‍👦 Ota-onangizning telefon raqamini kiriting:")
    await state.set_state(Registration.parent_phone)


@dp.message(Registration.parent_phone)
async def process_parent_phone(message: types.Message, state: FSMContext):
    await state.update_data(parent_phone=message.text)
    kb = InlineKeyboardBuilder()
    for idx, school in enumerate(ANGREN_SCHOOLS):
        kb.button(text=school, callback_data=f"school_{idx}")
    kb.adjust(4)
    await message.answer("🏫 Maktabingizni tanlang:", reply_markup=kb.as_markup())
    await state.set_state(Registration.school)


@dp.callback_query(Registration.school, F.data.startswith("school_"))
async def process_school(callback: types.CallbackQuery, state: FSMContext):
    idx = int(callback.data.replace("school_", ""))
    school_name = ANGREN_SCHOOLS[idx]
    await state.update_data(school=school_name)
    await callback.message.edit_text(f"🏫 Tanlandi: {school_name}")
    await callback.message.answer("🎓 Nechanchi sinfda o'qiysiz? (faqat raqam kiriting, masalan: 7)")
    await state.set_state(Registration.grade)
    await callback.answer()


@dp.message(Registration.grade)
async def process_grade(message: types.Message, state: FSMContext):
    if not message.text.strip().isdigit():
        await message.answer("Iltimos, faqat raqam kiriting (masalan: 7)")
        return
    await state.update_data(grade=message.text.strip())
    kb = InlineKeyboardBuilder()
    for letter in GRADE_LETTERS:
        kb.button(text=letter, callback_data=f"gradeletter_{letter}")
    kb.adjust(4)
    await message.answer("🔤 Sinf harfini tanlang:", reply_markup=kb.as_markup())
    await state.set_state(Registration.grade_letter)


@dp.callback_query(Registration.grade_letter, F.data.startswith("gradeletter_"))
async def process_grade_letter(callback: types.CallbackQuery, state: FSMContext):
    letter = callback.data.replace("gradeletter_", "")
    data = await state.get_data()
    full_grade = f"{data.get('grade')}-{letter}"
    await state.update_data(grade=full_grade)
    await callback.message.edit_text(f"🎓 Tanlandi: {full_grade}-sinf")
    kb = ReplyKeyboardBuilder()
    for filial in AVAILABLE_FILIALS:
        kb.button(text=filial)
    kb.adjust(2)
    await callback.message.answer("📍 Filialni tanlang:", reply_markup=kb.as_markup(resize_keyboard=True))
    await state.set_state(Registration.filial)
    await callback.answer()


@dp.message(Registration.filial)
async def process_filial(message: types.Message, state: FSMContext):
    if message.text not in AVAILABLE_FILIALS:
        await message.answer("Tugmalardan birini bosing!")
        return
    await state.update_data(filial=message.text, selected_courses=[])
    await state.set_state(Registration.subjects)
    await send_category_keyboard(message, state)


async def send_category_keyboard(target, state: FSMContext, edit: bool = False):
    data = await state.get_data()
    selected = data.get("selected_courses", [])
    kb = InlineKeyboardBuilder()
    for idx, (cat_name, subjects) in enumerate(COURSE_CATEGORY_LIST):
        count = sum(1 for s in subjects if s in selected)
        label = f"{cat_name}" + (f" ✅×{count}" if count else "")
        kb.button(text=label, callback_data=f"cat_{idx}")
    kb.button(text=f"➡️ Davom etish ({len(selected)} ta tanlangan)", callback_data="cat_done")
    kb.adjust(1)

    text = "📚 Kurs yo'nalishini tanlang:"
    if edit:
        await target.edit_text(text, reply_markup=kb.as_markup())
    else:
        await target.answer(text, reply_markup=kb.as_markup())


async def send_subject_list(message, state: FSMContext, cat_idx: int):
    data = await state.get_data()
    selected = data.get("selected_courses", [])
    cat_name, subjects = COURSE_CATEGORY_LIST[cat_idx]
    kb = InlineKeyboardBuilder()
    for i, subj in enumerate(subjects):
        mark = "✅ " if subj in selected else ""
        kb.button(text=f"{mark}{subj}", callback_data=f"csub_{cat_idx}_{i}")
    kb.button(text="⬅️ Kategoriyalarga qaytish", callback_data="cat_back")
    kb.adjust(1)

    if "Past o'zlashtiruvchi" in cat_name:
        header_text = (
            "🌱 <b>Past o'zlashtiruvchi o'quvchilar bilan ishlash bo'limi</b>\n\n"
            "💬 <i>Farzandingiz fanlarni o'zlashtirishda orqada qolyaptimi? Bizda yechim bor!</i>\n"
            "Har bir o'quvchi bilan alohida (individual) ishlab, uni iqtidorli darajagacha yetkazib beramiz!\n\n"
            "👇 <b>Dasturni tanlash uchun pastdagi tugmani bosing:</b>"
        )
    else:
        header_text = f"📌 <b>{cat_name}</b> yo'nalishidagi kurslarni tanlang:"

    await message.edit_text(header_text, reply_markup=kb.as_markup(), parse_mode="HTML")


@dp.callback_query(Registration.subjects, F.data.startswith("csub_"))
async def process_course_subject(callback: types.CallbackQuery, state: FSMContext):
    _, cat_idx_str, item_idx_str = callback.data.split("_")
    cat_idx = int(cat_idx_str)
    item_idx = int(item_idx_str)
    cat_name, subjects = COURSE_CATEGORY_LIST[cat_idx]
    subj = subjects[item_idx]

    data = await state.get_data()
    selected = data.get("selected_courses", [])
    if subj in selected:
        selected.remove(subj)
    else:
        selected.append(subj)
    await state.update_data(selected_courses=selected)
    await callback.answer()
    await send_subject_list(callback.message, state, cat_idx)


@dp.callback_query(Registration.subjects, F.data.startswith("cat_"))
async def process_category_nav(callback: types.CallbackQuery, state: FSMContext):
    action = callback.data.replace("cat_", "", 1)

    if action == "done":
        data = await state.get_data()
        selected_courses = data.get("selected_courses", [])
        if not selected_courses:
            await callback.answer("Kamida bitta fan tanlang!", show_alert=True)
            return
        kb = ReplyKeyboardBuilder()
        for t in AVAILABLE_TIMES:
            kb.button(text=t)
        kb.adjust(3)
        await callback.message.answer(
            "Dars vaqtini tanlang:", reply_markup=kb.as_markup(resize_keyboard=True)
        )
        await state.set_state(Registration.time_pref)
        await callback.answer()
        return

    if action == "back":
        await callback.answer()
        await send_category_keyboard(callback.message, state, edit=True)
        return

    cat_idx = int(action)
    await callback.answer()
    await send_subject_list(callback.message, state, cat_idx)


def escape_markdown(text):
    if text is None:
        return ""
    text = str(text)
    for ch in ("_", "*", "`", "["):
        text = text.replace(ch, "\\" + ch)
    return text


@dp.message(Registration.time_pref)
async def process_time_pref(message: types.Message, state: FSMContext):
    if message.text not in AVAILABLE_TIMES:
        await message.answer("Smenani tugmalardan tanlang!")
        return
    await state.update_data(time_pref=message.text)
    user_data = await state.get_data()

    try:
        save_to_excel(user_data, message.from_user.id)
    except Exception:
        logging.exception("Excel'ga yozishda xato:")

    asyncio.create_task(save_to_google_sheets(user_data, message.from_user.id))
    admin_id = os.getenv("ADMIN_ID")
    if admin_id:
        admin_text = (
            f"🆕 Yangi o'quvchi!\n"
            f"👤 {user_data.get('name')}\n"
            f"📞 {user_data.get('phone')}\n"
            f"👨‍👩‍👦 Ota-ona: {user_data.get('parent_phone')}\n"
            f"🏫 Maktab: {user_data.get('school')}, Sinf: {user_data.get('grade')}\n"
            f"📍 Filial: {user_data.get('filial')} | Smena: {user_data.get('time_pref')}\n"
            f"🕐 {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        )
        try:
            await bot.send_message(int(admin_id), admin_text)
        except Exception:
            logging.exception("Admin xabarida xato:")
    selected_courses = user_data.get("selected_courses", [])
    courses_output = "📚 Tanlangan kurslar:\n" + "".join(
        f"• {c.replace(chr(10), ' ')}\n" for c in selected_courses
    )

    student_report = (
        f"Muvaffaqiyatli royxatdan o'tdingiz!\n\n"
        f"O'quvchi: {escape_markdown(user_data.get('name'))}\n"
        f"Maktab/Sinf: {escape_markdown(user_data.get('school'))}, {escape_markdown(user_data.get('grade'))}\n"
        f"Filial: {escape_markdown(user_data.get('filial'))} | Smena: {escape_markdown(user_data.get('time_pref'))}\n\n"
        f"{courses_output}\n"
        f"+998 94 041 42 55\n+998 93 101 58 70"
    )

    shifokorlik_fanlari = COURSE_CATEGORIES.get("🩺 Shifokorlik yo'nalishi (5 ta fan bir joyda)", [])
    is_medical = any(course in shifokorlik_fanlari for course in selected_courses)
    if is_medical:
        try:
            photo = FSInputFile("IMG_20260619_235730_628.jpg")
            await message.answer_photo(
                photo=photo,
                caption=student_report,
                parse_mode="Markdown",
                reply_markup=get_main_menu()
            )
        except Exception:
            logging.exception("Rasm yuborishda xato:")
            await message.answer(
                text=student_report,
                parse_mode="Markdown",
                reply_markup=get_main_menu()
            )
    else:
        await message.answer(
            text=student_report,
            parse_mode="Markdown",
            reply_markup=get_main_menu()
        )
    tabrik_kb = InlineKeyboardBuilder()
    tabrik_kb.button(text="👥 Do'stimni taklif qilish", url="https://t.me/share/url?url=https://t.me/AngrenAkademiyaBot&text=Angren+Akademiyaga+qo%27shiling!")
    tabrik_kb.button(text="📍 Angren Akademiya kanali", url="https://t.me/AngrenAkademiya")
    tabrik_kb.adjust(1)
    await message.answer(
        f"🎉 Tabriklaymiz, {escape_markdown(user_data.get('name'))}!\n\n"
        f"🎓 Sizga sertifikat berildi!\n"
        f"Kelib Angren Akademiyadan olib keting!\n\n"
        f"📚 Darslarimizga ishtirok etib voucherlarni qo'lga kiriting!\n\n"
        f"🚀 O'z karyerangiz tomon — biz bilan qadam bosing!\n\n"
        f"🎁 Do'stlaringizni taklif qiling:\n"
        f"👉 1 ta do'st — 30 000 so'mlik voucher yoki sovg'a!\n"
        f"👉 2 ta do'st — 40 000 so'mlik voucher yoki sovg'a!\n"
        f"👉 3 va undan ko'p — 50 000 so'mlik voucher yoki sovg'alarni ham qo'lga kiriting!\n\n"
        f"⚠️ Joylar cheklangan — imkoniyatlarni qo'ldan boy bermang!",
        reply_markup=tabrik_kb.as_markup()
    )
    try:
        cert_path = generate_certificate(user_data.get("name"), message.from_user.id)
        await message.answer_photo(
            photo=FSInputFile(cert_path),
            caption="🎓 Tabriklaymiz! Sizning shaxsiy sertifikatingiz tayyor."
        )
        os.remove(cert_path)
    except Exception:
        logging.exception("Sertifikat generatsiyasida xato:")

    channel_id = os.getenv("CHANNEL_ID")
    if channel_id:
        try:
            await bot.send_message(int(channel_id),
                f"🆕 Yangi o'quvchi!\n"
                f"👤 {user_data.get('name')}\n"
                f"📞 {user_data.get('phone')}\n"
                f"📍 {user_data.get('filial')} | {user_data.get('time_pref')}\n"
                f"🕐 {datetime.now().strftime('%d.%m.%Y %H:%M')}"
            )
        except Exception:
            logging.exception("Kanalga xabar yuborishda xato:")

    await state.clear()
    grade_raw = user_data.get("grade", "")
    digits = "".join(ch for ch in grade_raw if ch.isdigit())
    if digits and 7 <= int(digits) <= 11:
        await start_diagnostic_test(message, state, int(digits))
@dp.message(F.text == "👤 Shaxsiy kabinet")
async def shaxsiy_kabinet(message: types.Message):
    student = await asyncio.to_thread(find_student_by_id, message.from_user.id)
    if not student:
        await message.answer("❌ Siz hali ro'yxatdan o'tmagansiz. Avval \"📝 Ro'yxatdan o'tish\" tugmasini bosing.")
        return
    await message.answer(
        f"👤 Sizning shaxsiy kabinetingiz\n\n"
        f"💳 Voucher balansiz: {student['voucher']:,} so'm\n"
        f"⚡️ Darslarni boshlashingiz bilan faollashadi!\n\n"
        f"👥 Taklif qilgan do'stlar: {student['referrals']} ta\n"
        f"🎓 Sertifikat: berildi ✅\n\n"
        f"🚀 O'z karyerangiz tomon — biz bilan qadam bosing!"
    )
@dp.callback_query(F.data == "profile")
async def profile_handler(callback: types.CallbackQuery):
    student = await asyncio.to_thread(find_student_by_id, callback.from_user.id)
    if not student:
        await callback.message.answer("❌ Siz hali ro'yxatdan o'tmagansiz. Avval \"📝 Ro'yxatdan o'tish\" tugmasini bosing.")
        return
    await callback.message.answer(
    f"👤 Sizning shaxsiy kabinetingiz\n\n"
    f"💳 Voucher balansiz: {student['voucher']:,} so'm\n"
    f"⚡️ Darslarni boshlashingiz bilan faollashadi!\n\n"
    f"👥 Taklif qilgan do'stlar: {student['referrals']} ta\n"
    f"🎓 Sertifikat: berildi ✅\n\n"
    f"🚀 O'z karyerangiz tomon — biz bilan qadam bosing!"
)



@dp.callback_query(F.data == "check_sub")
async def check_sub_callback(callback: types.CallbackQuery, state: FSMContext):
    if await check_subscription(callback.from_user.id):
        await callback.message.edit_text("✅ Rahmat! Endi botdan to'liq foydalanishingiz mumkin.")
        await callback.message.answer(
            "✨ Angren Akademiyasi rasmiy botiga xush kelibsiz!\n\n"
            "Kelajak akademiyasida o'z bilimingizni va farzandingiz kamolotini nazorat qiling.",
            reply_markup=get_main_menu()
        )
    else:
        await callback.answer("❌ Siz hali kanalga a'zo bo'lmagansiz!", show_alert=True)
        return
    await callback.answer()


@dp.callback_query(F.data == "cert")
async def cert_handler(callback: types.CallbackQuery):
    await callback.message.answer("🎓 Sertifikat yuklash bo'limi.")


async def handle_health(request):
    return web.Response(text="Angren Akademiya boti faol!")


async def start_web_server():
    app = web.Application()
    app.router.add_get("/", handle_health)
    runner = web.AppRunner(app)
    await runner.setup()
    port = int(os.getenv("PORT", 10000))
    site = web.TCPSite(runner, "0.0.0.0", port)
    await site.start()


async def pinger_loop():
    app_url = os.getenv("APP_URL")
    if not app_url:
        return
    await asyncio.sleep(10)
    while True:
        try:
            async with ClientSession() as session:
                async with session.get(app_url) as response:
                    pass
        except Exception:
            pass
        await asyncio.sleep(300)


async def main():
    await start_web_server()
    asyncio.create_task(pinger_loop())
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
